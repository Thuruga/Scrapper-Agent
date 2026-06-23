"""
Testes de contrato do endpoint POST /search/cross-marketplace/export.

Estrategia TDD (Wave 0 — RED scaffold):
  - Este arquivo descreve o comportamento ESPERADO do endpoint que ainda NAO
    existe. Todos os testes devem FALHAR neste momento, pois a rota, os modelos
    Pydantic e a funcao _sanitize_cell ainda nao foram implementados (Plan 02).
  - As falhas esperadas sao: 404 (rota ausente) ou ImportError (_sanitize_cell
    ausente) — NAO erros de colecao/sintaxe.
  - TestClient sincrono (fastapi.testclient.TestClient) importa a app de app.py.
  - verify_api_key sobrescrito via app.dependency_overrides para aceitar qualquer
    chave (os testes de happy path nao testam auth).
  - Resposta .content lida como io.BytesIO e validada com openpyxl.load_workbook().
  - TestSanitizeHelper testa _sanitize_cell como funcao pura; vai falhar com
    ImportError ate que Plan 02 implemente e exporte a funcao.

Requisitos cobertos: EXPORT-04, EXPORT-05, EXPORT-06, T-24-01, T-24-02, T-24-03, T-24-04.
"""
import io
import re

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app import app
from api.auth import verify_api_key

# ---------------------------------------------------------------------------
# Cabecalhos PT-BR esperados (contrato imutavel definido em CONTEXT.md)
# ---------------------------------------------------------------------------
EXPECTED_HEADERS = [
    "Plataforma",
    "Vendedor",
    "Título",
    "Preço",
    "Frete",
    "Preço Total",
    "Frete Grátis",
    "Score de Match",
    "Similar",
    "URL",
]

# ---------------------------------------------------------------------------
# Configuracao do cliente de teste — auth bypassada para todos exceto test_auth
# ---------------------------------------------------------------------------
app.dependency_overrides[verify_api_key] = lambda: "test-key"
client = TestClient(app)

# ---------------------------------------------------------------------------
# Fixture base — item com todos os campos necessarios
# ---------------------------------------------------------------------------
ITEM_BASE = {
    "marketplace": "Mercado Livre",
    "seller": "Vendedor Teste",
    "title": "Polo Piquet Aramis",
    "price": 199.90,
    "shipping_price": 15.00,
    "landed_price": 214.90,
    "is_free_shipping": False,
    "final_match_score": 87.4,
    "match_score": 72.0,
    "is_similar": False,
    "url": "https://example.com/produto",
    "_display_order": 0,
}

ENDPOINT = "/search/cross-marketplace/export"


# ===========================================================================
# TestExportEndpoint — contrato completo do endpoint
# ===========================================================================
class TestExportEndpoint:
    """
    Todos os testes desta classe devem falhar com 404 (rota nao existe) no
    Wave 0 / Plan 01. A rota sera implementada em Plan 02.
    """

    def test_happy_path(self):
        """
        POST com um item valido deve retornar 200.
        Planilha: sheet 'Busca SKU'; cabecalho na linha 1 com 10 colunas PT-BR.
        """
        response = client.post(
            ENDPOINT,
            json={"items": [ITEM_BASE], "target_sku": "ML.05.0326046"},
            headers={"X-API-Key": "test-key"},
        )
        assert response.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Busca SKU"]
        headers = [ws.cell(1, c).value for c in range(1, 11)]
        assert headers == EXPECTED_HEADERS

    def test_null_shipping(self):
        """
        Quando shipping_price=None e is_free_shipping=False:
          - Coluna Frete (col 5) deve ser 'A calcular'
          - Coluna Preco Total (col 6) deve ser igual ao preco do item (nunca 0)
        """
        item = {
            **ITEM_BASE,
            "shipping_price": None,
            "is_free_shipping": False,
            "_display_order": 0,
        }
        response = client.post(
            ENDPOINT,
            json={"items": [item], "target_sku": "TEST"},
            headers={"X-API-Key": "test-key"},
        )
        assert response.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Busca SKU"]
        assert ws.cell(2, 5).value == "A calcular"  # Frete
        assert ws.cell(2, 6).value == item["price"]  # Preco Total = preco do produto

    def test_free_shipping(self):
        """
        Quando is_free_shipping=True:
          - Frete Gratis (col 7) deve ser 'Sim'
          - Frete (col 5) deve ser 0
          - Preco Total (col 6) deve ser landed_price
        """
        item = {
            **ITEM_BASE,
            "is_free_shipping": True,
            "shipping_price": 0.0,
            "landed_price": 199.90,
            "_display_order": 0,
        }
        response = client.post(
            ENDPOINT,
            json={"items": [item], "target_sku": "TEST"},
            headers={"X-API-Key": "test-key"},
        )
        assert response.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Busca SKU"]
        assert ws.cell(2, 7).value == "Sim"   # Frete Gratis
        assert ws.cell(2, 5).value == 0        # Frete = 0
        assert ws.cell(2, 6).value == item["landed_price"]  # Preco Total

    def test_boolean_mapping(self):
        """
        Campos booleanos devem renderizar como string 'Sim' ou 'Nao', nunca True/False.
        """
        item_sim = {**ITEM_BASE, "is_free_shipping": True, "is_similar": True}
        item_nao = {**ITEM_BASE, "is_free_shipping": False, "is_similar": False}

        for item, expected_frete_gratis, expected_similar in [
            (item_sim, "Sim", "Sim"),
            (item_nao, "Não", "Não"),
        ]:
            response = client.post(
                ENDPOINT,
                json={"items": [item], "target_sku": "TEST"},
                headers={"X-API-Key": "test-key"},
            )
            assert response.status_code == 200
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            ws = wb["Busca SKU"]
            frete_gratis_val = ws.cell(2, 7).value
            similar_val = ws.cell(2, 9).value
            assert frete_gratis_val == expected_frete_gratis, (
                f"Frete Gratis esperado '{expected_frete_gratis}', obtido '{frete_gratis_val}'"
            )
            assert similar_val == expected_similar, (
                f"Similar esperado '{expected_similar}', obtido '{similar_val}'"
            )
            assert frete_gratis_val not in (True, False)
            assert similar_val not in (True, False)

    def test_score_rounding(self):
        """
        final_match_score=87.4 -> Score de Match (col 8) deve ser inteiro 87.
        Quando final_match_score==0, deve usar match_score como fallback.
        """
        # Caso 1: final_match_score > 0 -> arredondar para inteiro
        item = {**ITEM_BASE, "final_match_score": 87.4, "match_score": 72.0}
        response = client.post(
            ENDPOINT,
            json={"items": [item], "target_sku": "TEST"},
            headers={"X-API-Key": "test-key"},
        )
        assert response.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Busca SKU"]
        score_val = ws.cell(2, 8).value
        assert score_val == 87, f"Score esperado 87, obtido {score_val}"
        assert isinstance(score_val, int), f"Score deve ser inteiro, obtido {type(score_val)}"

        # Caso 2: final_match_score==0 -> fallback para match_score
        item_fallback = {**ITEM_BASE, "final_match_score": 0, "match_score": 72.0}
        response2 = client.post(
            ENDPOINT,
            json={"items": [item_fallback], "target_sku": "TEST"},
            headers={"X-API-Key": "test-key"},
        )
        assert response2.status_code == 200
        wb2 = openpyxl.load_workbook(io.BytesIO(response2.content))
        ws2 = wb2["Busca SKU"]
        score_fallback = ws2.cell(2, 8).value
        assert score_fallback == 72, f"Score fallback esperado 72, obtido {score_fallback}"

    def test_formula_injection(self):
        """
        Celulas cujo valor comeca com '=', '+', '-' ou '@' devem ser sanitizadas
        com um apostrofo (') como prefixo para evitar injecao de formula no Excel.
        """
        dangerous_prefixes = ["=SUM(A1)", "+BAD", "-BAD", "@cmd"]

        for dangerous_val in dangerous_prefixes:
            item = {**ITEM_BASE, "seller": dangerous_val}
            response = client.post(
                ENDPOINT,
                json={"items": [item], "target_sku": "TEST"},
                headers={"X-API-Key": "test-key"},
            )
            assert response.status_code == 200
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            ws = wb["Busca SKU"]
            cell_value = ws.cell(2, 2).value  # Vendedor = col 2
            assert cell_value is not None
            assert cell_value.startswith("'"), (
                f"Valor '{dangerous_val}' deveria ser sanitizado com apostrofo, "
                f"obtido: '{cell_value}'"
            )

    def test_formula_injection_url(self):
        """
        A coluna URL (col 10) tambem e string client-supplied e deve ser sanitizada.
        Um URL comecando com '=' deve receber apostrofo prefixo para evitar formula Excel.
        """
        item = {**ITEM_BASE, "url": '=HYPERLINK("https://evil.com","Click")'}
        response = client.post(
            ENDPOINT,
            json={"items": [item], "target_sku": "TEST"},
            headers={"X-API-Key": "test-key"},
        )
        assert response.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Busca SKU"]
        cell_value = ws.cell(2, 10).value  # URL = col 10
        assert cell_value is not None
        assert cell_value.startswith("'"), (
            f"URL com formula deveria ser sanitizado com apostrofo, obtido: '{cell_value}'"
        )

    def test_display_order(self):
        """
        Tres itens enviados em ordem embaralhada (_display_order 2, 0, 1) devem
        aparecer nas linhas da planilha em ordem crescente de _display_order (0, 1, 2).
        """
        items = [
            {**ITEM_BASE, "seller": "Vendedor C", "url": "https://example.com/c", "_display_order": 2},
            {**ITEM_BASE, "seller": "Vendedor A", "url": "https://example.com/a", "_display_order": 0},
            {**ITEM_BASE, "seller": "Vendedor B", "url": "https://example.com/b", "_display_order": 1},
        ]
        response = client.post(
            ENDPOINT,
            json={"items": items, "target_sku": "TEST"},
            headers={"X-API-Key": "test-key"},
        )
        assert response.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Busca SKU"]
        # Linha 1 = cabecalhos; linhas 2, 3, 4 = dados
        row2_seller = ws.cell(2, 2).value  # Vendedor A (_display_order=0)
        row3_seller = ws.cell(3, 2).value  # Vendedor B (_display_order=1)
        row4_seller = ws.cell(4, 2).value  # Vendedor C (_display_order=2)
        assert row2_seller == "Vendedor A", f"Linha 2 esperada 'Vendedor A', obtida '{row2_seller}'"
        assert row3_seller == "Vendedor B", f"Linha 3 esperada 'Vendedor B', obtida '{row3_seller}'"
        assert row4_seller == "Vendedor C", f"Linha 4 esperada 'Vendedor C', obtida '{row4_seller}'"

    def test_fidelity(self):
        """
        Todos os campos escalares enviados (marketplace, seller, title, price, url)
        devem aparecer identicos nas celulas correspondentes da planilha.
        O backend nao deve recomputar nem transformar os valores enviados.
        """
        item = {
            **ITEM_BASE,
            "marketplace": "Amazon",
            "seller": "Loja Fidelidade",
            "title": "Camiseta Premium Fidelidade",
            "price": 149.99,
            "url": "https://amazon.com.br/produto-fidelidade",
        }
        response = client.post(
            ENDPOINT,
            json={"items": [item], "target_sku": "TEST"},
            headers={"X-API-Key": "test-key"},
        )
        assert response.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["Busca SKU"]
        assert ws.cell(2, 1).value == item["marketplace"]  # Plataforma
        assert ws.cell(2, 2).value == item["seller"]        # Vendedor
        assert ws.cell(2, 3).value == item["title"]         # Titulo
        assert ws.cell(2, 4).value == item["price"]         # Preco
        assert ws.cell(2, 10).value == item["url"]          # URL

    def test_filename(self):
        """
        O header Content-Disposition deve conter um filename que:
          - Comeca com 'busca_sku_'
          - Contem um token derivado de search_query (ou target_sku como fallback)
          - Termina com _YYYYMMDD_HHMMSS.xlsx
          - Regex: ^busca_sku_.+_\\d{8}_\\d{6}\\.xlsx$
        """
        response = client.post(
            ENDPOINT,
            json={
                "items": [ITEM_BASE],
                "target_sku": "ML.05.0326046",
                "search_query": "polo piquet",
            },
            headers={"X-API-Key": "test-key"},
        )
        assert response.status_code == 200
        content_disposition = response.headers.get("content-disposition", "")
        assert content_disposition, "Header Content-Disposition ausente"
        # Extrair filename= do header
        match = re.search(r'filename="?([^";\s]+)"?', content_disposition)
        assert match, f"filename nao encontrado em Content-Disposition: {content_disposition}"
        filename = match.group(1)
        pattern = r"^busca_sku_.+_\d{8}_\d{6}\.xlsx$"
        assert re.match(pattern, filename), (
            f"Filename '{filename}' nao bate com padrao '{pattern}'"
        )

    def test_empty_items(self):
        """
        POST com items=[] deve retornar 400 ou 422.
        400 = backend rejeita explicitamente; 422 = Pydantic min_length=1.
        Ambos sao validos per CONTEXT decision.
        """
        response = client.post(
            ENDPOINT,
            json={"items": [], "target_sku": "TEST"},
            headers={"X-API-Key": "test-key"},
        )
        assert response.status_code in (400, 422), (
            f"Status esperado 400 ou 422 para items=[], obtido {response.status_code}"
        )

    def test_oversized_payload(self):
        """
        POST com 501 itens deve retornar 422 (Pydantic max_length=500).
        """
        items = [{**ITEM_BASE, "url": f"https://example.com/{i}", "_display_order": i}
                 for i in range(501)]
        response = client.post(
            ENDPOINT,
            json={"items": items, "target_sku": "TEST"},
            headers={"X-API-Key": "test-key"},
        )
        assert response.status_code == 422, (
            f"Status esperado 422 para 501 itens, obtido {response.status_code}"
        )

    def test_auth(self):
        """
        POST sem X-API-Key deve retornar 403 (verify_api_key herdado do api_router).
        Este teste temporariamente remove o override de auth e o restaura ao final.
        """
        # Remover override para testar autenticacao real
        app.dependency_overrides.pop(verify_api_key, None)
        try:
            response = client.post(
                ENDPOINT,
                json={"items": [ITEM_BASE], "target_sku": "TEST"},
                # Sem X-API-Key
            )
            # FastAPI returns 422 when the required Header(...) field is absent (validation
            # error before the dependency body runs), or 403 when the key is present but
            # invalid. Both indicate access denied — accept either per actual FastAPI behaviour.
            assert response.status_code in (403, 422), (
                f"Status esperado 403 ou 422 sem X-API-Key, obtido {response.status_code}"
            )
        finally:
            # Sempre restaurar o override para nao contaminar outros testes
            app.dependency_overrides[verify_api_key] = lambda: "test-key"


# ===========================================================================
# TestSanitizeHelper — testa _sanitize_cell como funcao pura (Plan 02 exporta)
# ===========================================================================
class TestSanitizeHelper:
    """
    Testa _sanitize_cell diretamente como funcao pura.
    Vai falhar com ImportError ate que Plan 02 implemente e exporte a funcao
    em api.routes_search. Isso e o comportamento esperado no Wave 0 RED.
    """

    def test_sanitize_formula_injection(self):
        """
        Strings comecando com '=', '+', '-', '@' devem receber um apostrofo prefixo.
        """
        from api.routes_search import _sanitize_cell  # noqa: PLC0415 — falha esperada no RED

        assert _sanitize_cell("=SUM(A1)") == "'=SUM(A1)"
        assert _sanitize_cell("+CMD") == "'+CMD"
        assert _sanitize_cell("-1+1") == "'-1+1"
        assert _sanitize_cell("@user") == "'@user"

    def test_sanitize_safe_string_unchanged(self):
        """
        Strings que nao comecam com caracteres perigosos devem passar sem alteracao.
        """
        from api.routes_search import _sanitize_cell  # noqa: PLC0415 — falha esperada no RED

        assert _sanitize_cell("Polo Piquet Aramis") == "Polo Piquet Aramis"
        assert _sanitize_cell("Mercado Livre") == "Mercado Livre"
        assert _sanitize_cell("") == ""
