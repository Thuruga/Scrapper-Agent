from __future__ import annotations

import io
from types import SimpleNamespace

import openpyxl
from fastapi.testclient import TestClient

from app import app
from core.models import BrandSearchResult, MapRule, PromotionInfo, SearchProductResult


def _patch_search_basics(monkeypatch, routes_search, rules):
    monkeypatch.setattr(
        routes_search.brand_service,
        "list_brands",
        lambda active_only=True: [SimpleNamespace(brand_key="aramis")],
    )
    monkeypatch.setattr(
        routes_search.map_rules_service,
        "list_rules",
        lambda active_only=True: rules,
    )
    monkeypatch.setattr(routes_search.search_history_service, "create_job", lambda **kwargs: None)
    monkeypatch.setattr(routes_search.search_history_service, "update_job", lambda **kwargs: None)


def _phase43_brand_result():
    return [
        BrandSearchResult(
            brand_key="aramis",
            brand_name="Aramis",
            products=[
                SearchProductResult(
                    brand="Aramis",
                    product_name="Polo Piquet",
                    url="https://www.aramis.com.br/p/polo",
                    price_full=300.0,
                    price_discount=250.0,
                    seller=None,
                    category="Polos",
                    promotions=[
                        PromotionInfo(
                            type="percentage_discount",
                            raw_text="17% OFF",
                            value=17,
                            unit="percent",
                        )
                    ],
                )
            ],
        )
    ]


def test_search_response_surfaces_map_and_promotions(monkeypatch):
    import api.routes_search as routes_search

    _patch_search_basics(
        monkeypatch,
        routes_search,
        [MapRule(id="rule-brand", scope="brand", target="Aramis", min_price=275.0)],
    )

    async def fake_search_all_brands(**kwargs):
        return _phase43_brand_result()

    monkeypatch.setattr(routes_search.engine_factory, "search_all_brands", fake_search_all_brands)

    response = TestClient(app).post(
        "/search",
        headers={"X-API-Key": "dev-api-key"},
        json={"query": "polo", "brands": ["aramis"], "max_per_brand": 1},
    )

    assert response.status_code == 200
    product = response.json()["results"][0]["products"][0]
    assert product["map_violation"] is True
    assert product["map_price_floor"] == 275.0
    assert product["map_rule_scope"] == "brand"
    assert product["map_rule_id"] == "rule-brand"
    assert product["map_infractor"] == "Aramis"
    assert product["promotions"][0]["type"] == "percentage_discount"
    assert product["promotions"][0]["raw_text"].endswith("OFF")


def test_cross_marketplace_rows_can_include_map_metadata(monkeypatch):
    import services.cross_marketplace_service as cross_module

    monkeypatch.setattr(
        cross_module.map_rules_service,
        "list_rules",
        lambda active_only=True: [
            MapRule(id="rule-cross", scope="brand", target="aramis", min_price=250.0)
        ],
    )

    rows = [
        {
            "marketplace": "Mercado Livre",
            "seller": "Loja Parceira",
            "title": "Polo Piquet Aramis",
            "price": 199.0,
            "url": "https://produto.mercadolivre.com.br/polo",
        }
    ]
    cross_module.cross_marketplace_service._apply_phase43_metadata(rows, "aramis")

    assert rows[0]["map_violation"] is True
    assert rows[0]["map_rule_id"] == "rule-cross"
    assert rows[0]["map_infractor"] == "Loja Parceira"
    assert rows[0]["map_infractor_is_default"] is False
    assert rows[0]["promotions"] == []


def test_search_export_adds_phase43_columns_only_when_present(monkeypatch):
    import api.routes_search as routes_search

    captured: dict[str, object] = {}
    _patch_search_basics(
        monkeypatch,
        routes_search,
        [MapRule(id="rule-export", scope="brand", target="Aramis", min_price=275.0)],
    )

    async def fake_search_all_brands(**kwargs):
        return _phase43_brand_result()

    class _FakeEngine:
        async def get_product_details(self, product_url: str):
            return None

    class _FakeExcelWriter:
        def __init__(self, *args, **kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

    def fake_to_excel(self, *args, **kwargs):
        captured["columns"] = self.columns.tolist()
        captured["row"] = self.iloc[0].to_dict()

    monkeypatch.setattr(routes_search.engine_factory, "search_all_brands", fake_search_all_brands)
    monkeypatch.setattr(routes_search.engine_factory, "get_engine", lambda brand_key: _FakeEngine())
    monkeypatch.setattr(routes_search.pd, "ExcelWriter", _FakeExcelWriter)
    monkeypatch.setattr(routes_search.pd.DataFrame, "to_excel", fake_to_excel)

    response = TestClient(app).post(
        "/search/export",
        headers={"X-API-Key": "dev-api-key"},
        json={"query": "polo", "brands": ["aramis"], "max_per_brand": 1},
    )

    assert response.status_code == 200
    assert "map_violation" in captured["columns"]
    assert "promotions" in captured["columns"]
    assert captured["row"]["map_rule_id"] == "rule-export"
    assert "percentage_discount" in captured["row"]["promotions"]


def test_cross_export_serializes_phase43_columns():
    item = {
        "marketplace": "Mercado Livre",
        "seller": "Loja Parceira",
        "title": "Polo Piquet Aramis",
        "price": 199.0,
        "shipping_price": None,
        "landed_price": 199.0,
        "is_free_shipping": False,
        "final_match_score": 90.0,
        "match_score": 90.0,
        "is_similar": False,
        "url": "https://produto.mercadolivre.com.br/polo",
        "map_violation": True,
        "map_price_floor": 250.0,
        "map_rule_scope": "brand",
        "map_rule_id": "rule-cross",
        "map_infractor": "Loja Parceira",
        "promotions": [{"type": "generic_badge", "raw_text": "Oferta relampago"}],
    }

    response = TestClient(app).post(
        "/search/cross-marketplace/export",
        headers={"X-API-Key": "dev-api-key"},
        json={"items": [item], "target_sku": "ML.05.0326046"},
    )

    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb["Busca SKU"]
    headers = [cell.value for cell in ws[1]]
    assert "Violacao MAP" in headers
    assert "Promocoes" in headers
    row = {header: ws.cell(2, idx + 1).value for idx, header in enumerate(headers)}
    assert row["Violacao MAP"] == "Sim"
    assert row["Piso MAP"] == 250.0
    assert "Oferta relampago" in row["Promocoes"]
